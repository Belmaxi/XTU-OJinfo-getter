# -*- coding:utf-8 -*-
# @Author : Belmaxi

#自动安装依赖
try:
    print("正在检查依赖")
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    import os
except ImportError:
    import pip
    pip.main(["install", "requests", "bs4", "pandas"])
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    import os

# 解析url,返回html文本
def get_single_html(url):
    headers = {
        "cookie": "PHPSESSID=afpg8s9ni96hs03p0jmde9lgb7; think_template=default",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.76"
    }
    resp = requests.get(url, headers=headers)
    if resp.status_code == 200:
        return resp.text
    else:
        return None


# 将一行信息解析，返回二维列表，[[”“,info],[]]
# 排名 姓名 学号 班级 题单...
def parse_info(info, header):
    res = list(map(lambda x: [x], header))

    _id = (info[1].find("span", style="font-size:9px").get_text())[1:13]
    name = info[1].find("span", style="font-size:13px;color:#FF851B").get_text()
    _class = (info[1].find("span", style="font-size:9px").get_text())[13:]

    res[0] = ["排名", info[0].get_text()]
    res[1].append(name)
    res[2].append(_id)
    res[3].append(_class)

    for i in range(4,len(header)):
        res[i].append(info[i-2].get_text())

    return res


def get_all_standing(url):
    html = get_single_html(url)
    soup = BeautifulSoup(html, "html.parser")
    standing = (
        soup.find("div", id="content")
        .find("table", id="standing")
        .find_all("tr")
    )
    head = standing[0].find_all("th")
    # 使用原表头
    lst = list(map(lambda x: x.get_text(), head))

    del lst[1]
    lst.insert(1, "班级")
    lst.insert(1, "学号")
    lst.insert(1, "姓名")

    excel_info = []
    for std in standing[1:]:
        info = std.find_all("td")
        parsed_info = parse_info(info, lst)
        if parsed_info[2][1][0:4] != "2023":
            continue
        dic = {}
        for i in parsed_info:
            dic[i[0]] = i[1]

        excel_info.append(dic)
    return excel_info


url = "https://acm.xtu.edu.cn/exam/index.php/solution/stat/cids/409,410,411,412,413,414,415,416,417,418,419,420,421,422,423,424,425,426,427,428,429,430"
urls = [
    [
        f"https://acm.xtu.edu.cn/exam/index.php/solution/stat/order/{i[0]}/cids/409,410,411,412,413,414,415,416,417,418,419,420,421,422,423,424,425,426,427,428,429,430",
        i[1]] for i in [[1, "班级"], [2, "专业"], [3, "学号"]]
]


print("========正在爬取总表=======")
res = [["总数", pd.DataFrame(get_all_standing(url))]]
for i in urls:
    print(f"========正在爬取{i[1]}顺序表=======")
    res.append([i[1], pd.DataFrame(get_all_standing(i[0]))])

print(f"======================")
print("正在导出excel")

with pd.ExcelWriter('standing.xlsx') as writer:
    for i in res:
        i[1].to_excel(writer, sheet_name=i[0], index=False)

wb = load_workbook('standing.xlsx')

for ws in wb._sheets:
    ws.column_dimensions['C'].width = 20.0  # 调整列A宽
    ws.column_dimensions['D'].width = 30.0

wb.save("standing.xlsx")

print("导出完成")

print("================================")
input("press any to continue...")